from django.shortcuts import render
from custom.models import Page, Section
from main.models import SiteContent
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render, redirect
from django.contrib import messages
from main.models import Place, Institution
from django.contrib.admin.views.decorators import staff_member_required
from django.views.generic import TemplateView
from django.db.models import Count
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import render
import matplotlib
matplotlib.use('Agg')  # Устанавливаем бэкенд перед импортом pyplot
import matplotlib.pyplot as plt
import io
import base64
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Anketa
from datetime import datetime, timedelta
from django.utils.timezone import make_aware


from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse



@method_decorator(login_required, name='dispatch')
class BaseStatisticsView(TemplateView):
    template_name = "admin/statistics.html"
    group_by_field = None
    title = "Статистика"

    def get(self, request, *args, **kwargs):
        # Проверяем наличие параметра `export=excel` в GET-запросе
        if request.GET.get('export') == 'excel':
            start_date = request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
            end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

            # Преобразуем end_date в конец дня
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            end_date = make_aware(end_date)
            queryset = Anketa.objects.filter(created_at__range=[start_date, end_date])
            statistics = queryset.values(f'{self.group_by_field}__name').annotate(count=Count('id')).order_by('-count')
            return self.export_to_excel(statistics, start_date, end_date)

        # Если экспорт не запрашивается, вызываем стандартный метод `get()`
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        start_date = self.request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

        # Преобразуем end_date в конец дня
        end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        end_date = make_aware(end_date)
        show_top_10 = self.request.GET.get('show_top_10') == 'on'

        queryset = Anketa.objects.filter(created_at__range=[start_date, end_date])
        statistics = queryset.values(f'{self.group_by_field}__name').annotate(count=Count('id')).order_by('-count')

        total_count = queryset.count()
        others_count = 0

        chart_base64 = None
        if show_top_10 and len(statistics) > 10:
            top_10 = list(statistics[:10])
            others_count = sum(item['count'] for item in statistics[10:])
            max_count = max(item['count'] for item in top_10)

            if others_count <= 2 * max_count:
                top_10.append({f'{self.group_by_field}__name': 'Прочие', 'count': others_count})

            statistics = top_10

            labels = [item[f'{self.group_by_field}__name'] for item in statistics]
            counts = [item['count'] for item in statistics]

            fig, ax = plt.subplots(figsize=(6, 4))
            ax.pie(counts, labels=labels, autopct='%1.1f%%', startangle=140, colors=plt.cm.Paired.colors, textprops={'fontsize': 8})
            ax.set_title(f'Распределение заявок по {self.title} (Топ-10)', fontsize=10)

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight')
            plt.close(fig)
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            chart_base64 = f"data:image/png;base64,{image_base64}"

        context.update({
            'statistics': statistics,
            'total_count': total_count,
            'others_count': others_count,
            'start_date': start_date,
            'end_date': end_date.strftime('%Y-%m-%d'),
            'show_top_10': show_top_10,
            'chart_base64': chart_base64,
            'title': self.title,
        })
        return context

    def export_to_excel(self, statistics, start_date, end_date):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Статистика"

        headers = [self.title, "Количество заявок"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header

        for row_num, item in enumerate(statistics, 2):
            worksheet[f"A{row_num}"] = item[f'{self.group_by_field}__name']
            worksheet[f"B{row_num}"] = item['count']

        worksheet[f"A{row_num + 1}"] = "Всего"
        worksheet[f"B{row_num + 1}"] = sum(item['count'] for item in statistics)

        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 20

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="statistics_{start_date}_{end_date}.xlsx"'
        workbook.save(response)
        return response




class StatisticsByInstitutionView(BaseStatisticsView):
    group_by_field = 'institution'
    title = "Учреждение"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        start_date = self.request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

        # Преобразуем end_date в конец дня
        end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        end_date = make_aware(end_date)
        show_top_10 = self.request.GET.get('show_top_10') == 'on'

        queryset = Anketa.objects.filter(
            created_at__range=[start_date, end_date],
            status__in=['processed', 'feedback_received']
        )
        statistics = queryset.values(f'{self.group_by_field}__name').annotate(count=Count('id')).order_by('-count')

        total_count = queryset.count()
        others_count = 0

        chart_base64 = None
        if show_top_10 and len(statistics) > 10:
            top_10 = list(statistics[:10])
            others_count = sum(item['count'] for item in statistics[10:])
            max_count = max(item['count'] for item in top_10)

            if others_count <= 2 * max_count:
                top_10.append({f'{self.group_by_field}__name': 'Прочие', 'count': others_count})

            statistics = top_10

            labels = [item[f'{self.group_by_field}__name'] for item in statistics]
            counts = [item['count'] for item in statistics]

            fig, ax = plt.subplots(figsize=(6, 4))
            ax.pie(counts, labels=labels, autopct='%1.1f%%', startangle=140, colors=plt.cm.Paired.colors, textprops={'fontsize': 8})
            ax.set_title(f'Распределение заявок по {self.title} (Топ-10)', fontsize=10)

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight')
            plt.close(fig)
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            chart_base64 = f"data:image/png;base64,{image_base64}"

        context.update({
            'statistics': statistics,
            'total_count': total_count,
            'others_count': others_count,
            'start_date': start_date,
            'end_date': end_date.strftime('%Y-%m-%d'),
            'show_top_10': show_top_10,
            'chart_base64': chart_base64,
            'title': self.title,
        })
        return context

class StatisticsByCityView(BaseStatisticsView):
    group_by_field = 'city'
    title = "Город"




def get_relation_statistics(queryset):
    """
    Возвращает статистику по степеням родства.
    """
    relation_choices = dict(Anketa.RELATION_CHOICES)
    statistics = queryset.values('relation').annotate(count=Count('id')).order_by('-count')

    # Преобразуем ключи в их человекочитаемые значения
    formatted_statistics = [
        {'relation': relation_choices[item['relation']], 'count': item['count']}
        for item in statistics
    ]

    return formatted_statistics

class StatisticsByRelationView(TemplateView):
    template_name = "admin/statistics.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        start_date = self.request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

        # Преобразуем end_date в конец дня
        end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        end_date = make_aware(end_date)

        queryset = Anketa.objects.filter(created_at__range=[start_date, end_date])
        statistics = queryset.values('relation').annotate(count=Count('id')).order_by('-count')

        # Преобразуем ключи в человекочитаемые значения
        relation_choices = dict(Anketa.RELATION_CHOICES)
        formatted_statistics = [
            {'relation': relation_choices.get(item['relation'], "Не указано"), 'count': item['count']}
            for item in statistics
        ]

        total_count = sum(item['count'] for item in statistics)

        context.update({
            'statistics': formatted_statistics,
            'total_count': total_count,
            'title': "Статистика по степеням родства",
            'start_date': start_date,
            'end_date': end_date.strftime('%Y-%m-%d'),
        })
        return context
    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            start_date = request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
            end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

            # Преобразуем end_date в конец дня
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            end_date = make_aware(end_date)

            queryset = Anketa.objects.filter(created_at__range=[start_date, end_date])
            statistics = get_relation_statistics(queryset)

            # Экспортируем данные
            return self.export_to_excel(statistics, start_date, end_date)

        return super().get(request, *args, **kwargs)
    def export_to_excel(self, statistics, start_date, end_date):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Статистика по степеням родства"

        headers = ["Степень родства", "Количество заявок"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header

        for row_num, item in enumerate(statistics, 2):
            worksheet[f"A{row_num}"] = item['relation']
            worksheet[f"B{row_num}"] = item['count']

        worksheet[f"A{row_num + 1}"] = "Всего"
        worksheet[f"B{row_num + 1}"] = sum(item['count'] for item in statistics)

        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 20

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="relation_statistics_{start_date}_{end_date}.xlsx"'
        workbook.save(response)
        return response

class StatisticsByContactView(TemplateView):
    template_name = "admin/statistics_by_contact.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Определяем тип статистики на основе имени маршрута
        if self.request.resolver_match.url_name == 'statistics_by_contact':
            stat_type = 'contact'
        elif self.request.resolver_match.url_name == 'statistics_by_time':
            stat_type = 'time'
        elif self.request.resolver_match.url_name == 'statistics_by_reasons':
            stat_type = 'reasons'
        elif self.request.resolver_match.url_name == 'statistics_by_sources':
            stat_type = 'sources'
        else:
            stat_type = 'unknown'

        # Параметры фильтрации
        start_date = self.request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

        end_date_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        end_date_aware = make_aware(end_date_dt)

        queryset = Anketa.objects.filter(created_at__range=[start_date, end_date_aware])

        # Обработка разных типов статистики
        if stat_type == 'contact':
            statistics = self.get_contact_statistics(queryset)
            title = "Статистика по предпочитаемым способам связи"
            choices = dict(Anketa.CONTACT_CHOICES)
            field_name = 'contact'
        elif stat_type == 'time':
            statistics = self.get_time_statistics(queryset)
            title = "Статистика по предпочитаемому времени связи"
            choices = dict(Anketa.TIME_CHOICES)
            statistics = [{'time': item['preferred_time'], 'count': item['count']} for item in statistics]
            field_name = 'time'
        elif stat_type == 'reasons':
            statistics = self.get_reasons_statistics(queryset)
            title = "Статистика по причинам обращения"
            choices = dict(Anketa.REASON_CHOICES)
            field_name = 'reason'
        elif stat_type == 'sources':
            statistics = self.get_sources_statistics(queryset)
            title = "Статистика по источникам информации"
            choices = dict(Anketa.SOURCE_CHOICES)
            field_name = 'source'
        else:
            statistics = []
            title = "Неизвестная статистика"
            choices = {}
            field_name = 'unknown'

        # Форматируем статистику
        formatted_statistics = [
            {'label': choices.get(item[field_name], "Не указано"), 'count': item['count']}
            for item in statistics
        ]

        total_count = sum(item['count'] for item in statistics)

        context.update({
            'statistics': formatted_statistics,
            'total_count': total_count,
            'title': title,
            'start_date': start_date,
            'end_date': end_date,
        })
        return context

    def get_contact_statistics(self, queryset):
        contact_counts = {}
        for anketa in queryset:
            contacts = anketa.preferred_contact
            if isinstance(contacts, list):
                for contact in contacts:
                    contact_counts[contact] = contact_counts.get(contact, 0) + 1
            elif isinstance(contacts, str):
                contact_counts[contacts] = contact_counts.get(contact, 0) + 1
        return [{'contact': contact, 'count': count} for contact, count in contact_counts.items()]

    def get_time_statistics(self, queryset):
        return queryset.values('preferred_time').annotate(count=Count('id')).order_by('-count')

    def get_reasons_statistics(self, queryset):
        reasons_counts = {}
        for anketa in queryset:
            reasons = anketa.reasons
            if isinstance(reasons, list):
                for reason in reasons:
                    reasons_counts[reason] = reasons_counts.get(reason, 0) + 1
        return [{'reason': reason, 'count': count} for reason, count in reasons_counts.items()]

    def get_sources_statistics(self, queryset):
        sources_counts = {}
        for anketa in queryset:
            sources = anketa.sources
            if isinstance(sources, list):
                for source in sources:
                    sources_counts[source] = sources_counts.get(source, 0) + 1
        return [{'source': source, 'count': count} for source, count in sources_counts.items()]


    
class StatisticsByAgeView(TemplateView):
    template_name = "admin/statistics_by_age.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получение параметров даты из GET-запроса или установка значений по умолчанию
        start_date = self.request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

        # Преобразуем end_date в конец дня
        end_date_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        end_date_aware = make_aware(end_date_dt)

        # Фильтрация записей в диапазоне дат
        queryset = Anketa.objects.filter(created_at__range=[start_date, end_date_aware])

        # Категории возрастов
        age_categories = [
            (0, 6, "0 - 6 месяцев"),
            (7, 12, "7 - 12 месяцев"),
            (13, 24, "13 - 24 месяца (2 года)"),
            (25, 48, "25 - 48 месяцев (2 - 4 года)"),
            (49, 84, "49 - 84 месяца (5 - 7 лет)"),
            (85, 1000, "Больше 7 лет"),
        ]

        # Подсчёт записей в каждой категории
        age_statistics = []
        for min_age, max_age, label in age_categories:
            count = queryset.filter(child_age_in_months__gte=min_age, child_age_in_months__lte=max_age).count()
            age_statistics.append({'label': label, 'count': count})

        # Общее количество записей
        total_count = sum(item['count'] for item in age_statistics)

        context.update({
            'statistics': age_statistics,
            'total_count': total_count,
            'title': "Статистика по возрасту ребёнка",
            'start_date': start_date,
            'end_date': end_date,
        })
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            start_date = request.GET.get('start_date', (timezone.now() - timezone.timedelta(days=7)).strftime('%Y-%m-%d'))
            end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))

            # Преобразуем end_date в конец дня
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            end_date_aware = make_aware(end_date_dt)

            queryset = Anketa.objects.filter(created_at__range=[start_date, end_date_aware])
            return self.export_to_excel(queryset, start_date, end_date)

        return super().get(request, *args, **kwargs)

    def export_to_excel(self, queryset, start_date, end_date):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Статистика по возрасту"

        # Заголовки таблицы
        headers = ["Возраст", "Количество заявок"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header
            worksheet[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)

        # Категории возрастов
        age_categories = [
            (0, 6, "0 - 6 месяцев"),
            (7, 12, "7 - 12 месяцев"),
            (13, 24, "13 - 24 месяца (2 года)"),
            (25, 48, "25 - 48 месяцев (2 - 4 года)"),
            (49, 84, "49 - 84 месяца (5 - 7 лет)"),
            (85, 1000, "Больше 7 лет"),
        ]

        # Заполняем таблицу
        row_num = 2
        for min_age, max_age, label in age_categories:
            count = queryset.filter(child_age_in_months__gte=min_age, child_age_in_months__lte=max_age).count()
            worksheet[f"A{row_num}"] = label
            worksheet[f"B{row_num}"] = count
            row_num += 1

        # Итоговая строка
        worksheet[f"A{row_num}"] = "Всего"
        worksheet[f"B{row_num}"] = sum(
            queryset.filter(child_age_in_months__gte=min_age, child_age_in_months__lte=max_age).count()
            for min_age, max_age, label in age_categories
        )
        worksheet[f"A{row_num}"].font = openpyxl.styles.Font(bold=True)
        worksheet[f"B{row_num}"].font = openpyxl.styles.Font(bold=True)

        # Настройка ширины колонок
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 20

        # Подготовка ответа
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename="age_statistics_{start_date}_{end_date}.xlsx"'
        workbook.save(response)
        return response





@csrf_exempt
def anketa_view(request):
    if request.method == 'POST':
        # Определяем, является ли запрос AJAX
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

        # Получаем данные из POST-запроса
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name')
        relation = request.POST.get('relation')
        relation_other = request.POST.get('relation_other')
        main_phone = request.POST.get('main_phone')
        additional_phone = request.POST.get('additional_phone')
        main_email = request.POST.get('main_email')
        additional_email = request.POST.get('additional_email')
        preferred_contact = request.POST.getlist('preferred_contact')
        preferred_time = request.POST.get('preferred_time')
        child_last_name = request.POST.get('child_last_name')
        child_first_name = request.POST.get('child_first_name')
        child_middle_name = request.POST.get('child_middle_name')
        child_birth_date = request.POST.get('child_birth_date')
        city_id = request.POST.get('city')
        city = Place.objects.get(id=city_id) if city_id else None
        street = request.POST.get('street')
        house = request.POST.get('house')
        apartment = request.POST.get('apartment')
        reasons = request.POST.getlist('reasons')
        reason_other = request.POST.get('reason_other')
        referral_document = request.FILES.get('referral_document')
        sources = request.POST.getlist('sources')
        source_other = request.POST.get('source_other')
        consent = request.POST.get('consent') == 'on'

        # Серверная валидация данных
        errors = []

        # Пример простой валидации
        if not last_name:
            errors.append('Фамилия обязательна.')
        if not first_name:
            errors.append('Имя обязательно.')
        if not main_phone:
            errors.append('Основной номер телефона обязателен.')
        if not main_email:
            errors.append('Основной электронный адрес обязателен.')
        if not child_last_name:
            errors.append('Фамилия ребенка обязательна.')
        if not child_first_name:
            errors.append('Имя ребенка обязательно.')
        if not child_birth_date:
            errors.append('Дата рождения ребенка обязательна.')
        if not street:
            errors.append('Улица/микрорайон обязательны.')
        if not house:
            errors.append('Дом обязателен.')

        # Проверка обязательных полей и других условий
        # Добавьте дополнительные проверки по необходимости

        if is_ajax:
            if errors:
                return JsonResponse({'success': False, 'errors': errors})
            try:
                # Сохранение данных в модель Anketa
                anketa = Anketa(
                    last_name=last_name,
                    first_name=first_name,
                    middle_name=middle_name,
                    relation=relation,
                    relation_other=relation_other,
                    main_phone=main_phone,
                    additional_phone=additional_phone,
                    main_email=main_email,
                    additional_email=additional_email,
                    preferred_contact=preferred_contact,
                    preferred_time=preferred_time,
                    child_last_name=child_last_name,
                    child_first_name=child_first_name,
                    child_middle_name=child_middle_name,
                    child_birth_date=child_birth_date,
                    city=city,
                    street=street,
                    house=house,
                    apartment=apartment,
                    reasons=reasons,
                    reason_other=reason_other,
                    referral_document=referral_document,
                    sources=sources,
                    source_other=source_other,
                    consent=consent
                )
                anketa.save()
                return JsonResponse({'success': True})
            except Exception as e:
                # Логирование ошибки может быть добавлено здесь
                return JsonResponse({'success': False, 'errors': ['Произошла ошибка при сохранении анкеты.']})
        else:
            # Обработка обычного POST-запроса без AJAX
            if errors:
                # Здесь вы можете добавить сообщения об ошибках, например, используя Django messages
                return redirect('anketa')  # Или другой подход по обработке ошибок
            try:
                anketa = Anketa(
                    last_name=last_name,
                    first_name=first_name,
                    middle_name=middle_name,
                    relation=relation,
                    relation_other=relation_other,
                    main_phone=main_phone,
                    additional_phone=additional_phone,
                    main_email=main_email,
                    additional_email=additional_email,
                    preferred_contact=preferred_contact,
                    preferred_time=preferred_time,
                    child_last_name=child_last_name,
                    child_first_name=child_first_name,
                    child_middle_name=child_middle_name,
                    child_birth_date=child_birth_date,
                    city=city,
                    street=street,
                    house=house,
                    apartment=apartment,
                    reasons=reasons,
                    reason_other=reason_other,
                    referral_document=referral_document,
                    sources=sources,
                    source_other=source_other,
                    consent=consent
                )
                anketa.save()
                return redirect('anketa')
            except Exception as e:
                # Логирование ошибки может быть добавлено здесь
                return redirect('anketa')  # Или другой подход по обработке ошибок

    # Обработка GET-запроса
    cities = Place.objects.all().order_by('name')

    context = {
        'title': 'Анкета на помощь',
        'menu_sections': Section.objects.all().order_by('order'),
        'account_section_id': int(SiteContent.objects.get(name='account_section_id').content),
        'menu_pages': Page.objects.filter(section=None),
        'rules_url': SiteContent.objects.get(name='rules_url').content,
        'header_content': [
            SiteContent.objects.get(name='email').content,
            SiteContent.objects.get(name='phone').content,
            SiteContent.objects.get(name='phone').content.translate(str.maketrans({' ': '', '-': '', '(': '', ')': ''}))
        ],
        'cities': cities,  # Передаём список городов в шаблон
    }

    return render(request, 'anketa/anketa.html', context)

